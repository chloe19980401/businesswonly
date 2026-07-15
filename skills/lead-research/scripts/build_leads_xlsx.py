#!/usr/bin/env python3
"""
Build a sourced leads workbook from per-country JSON files.

Usage:
    python3 build_leads_xlsx.py <leads_dir> <countries.json> <output.xlsx> "<Title>"

<leads_dir>       folder with <CODE>.json files (schema: company, country, city, category,
                  productFocus, website, email, phone, sourceUrl, confidence, notes).
                  Optional enrich/<CODE>.json: [{company, email, phone, contactSourceUrl}].
<countries.json>  {"<CODE>": {"name": "中文名", "region": "地区"}, ...} — order + labels.
<output.xlsx>     output path.
<Title>           summary-sheet title.

Produces: 汇总 (COUNTIF summary), 全部线索 (all rows), one sheet per country, 说明 (notes).
Emails/phones/websites/sources render as clickable links. Run the xlsx recalc.py afterwards.
"""
import json
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CAT = {
    'importer_distributor': '进口商/经销商', 'hardware_retail': '五金/建材零售',
    'security_integrator': '安防/门禁集成商', 'developer': '地产开发商',
    'contractor': '建筑总包', 'hotel_group': '酒店集团',
}
HEAD = ['公司名称', '国家', '地区', '城市', '客户类别', '主营 / 产品', '网站',
        '公开邮箱', '公开电话', '信息来源', '置信度', '备注']
WCOL = [30, 10, 12, 14, 16, 34, 30, 26, 20, 42, 8, 40]


def load(leads_dir, codes):
    merged = {}
    for c in codes:
        path = os.path.join(leads_dir, f'{c}.json')
        rows = json.load(open(path, encoding='utf-8')) if os.path.exists(path) else []
        enr = {}
        ep = os.path.join(leads_dir, 'enrich', f'{c}.json')
        if os.path.exists(ep):
            for e in json.load(open(ep, encoding='utf-8')):
                enr[e['company'].strip()] = e
        for r in rows:
            e = enr.get(r.get('company', '').strip())
            if e:
                if not r.get('email') and e.get('email'):
                    r['email'] = e['email']
                if not r.get('phone') and e.get('phone'):
                    r['phone'] = e['phone']
        merged[c] = rows
    return merged


def main():
    if len(sys.argv) < 5:
        print(__doc__)
        sys.exit(1)
    leads_dir, countries_path, out_path, title = sys.argv[1:5]
    countries = json.load(open(countries_path, encoding='utf-8'))
    codes = list(countries.keys())
    CN = {c: countries[c]['name'] for c in codes}
    REG = {c: countries[c].get('region', '') for c in codes}
    merged = load(leads_dir, codes)

    navy = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    hdr = PatternFill('solid', fgColor='1F3A5F')
    thin = Side(style='thin', color='D9DEE7')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical='top', wrap_text=True)
    ctr = Alignment(horizontal='center', vertical='top')
    lf = Font(name='Arial', color='1768E5', underline='single', size=9)
    cf = Font(name='Arial', size=9)

    def hdrst(ws):
        for j, h in enumerate(HEAD, 1):
            c = ws.cell(1, j, h)
            c.font = navy; c.fill = hdr; c.border = bd
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for j, w in enumerate(WCOL, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = 'A2'

    def wr(ws, i, row, code):
        vals = [row.get('company', ''), CN.get(code, code), REG.get(code, ''),
                row.get('city', ''), CAT.get(row.get('category', ''), row.get('category', '')),
                row.get('productFocus', ''), row.get('website', ''), row.get('email', ''),
                row.get('phone', ''), row.get('sourceUrl', ''),
                '高' if row.get('confidence') == 'high' else '中', row.get('notes', '')]
        for j, v in enumerate(vals, 1):
            c = ws.cell(i, j, v); c.alignment = wrap; c.border = bd; c.font = cf
            if j == 7 and str(v).startswith('http'):
                c.hyperlink = v; c.font = lf
            if j == 8 and '@' in str(v):
                c.hyperlink = 'mailto:' + v; c.font = lf
            if j == 10 and str(v).startswith('http'):
                c.hyperlink = v; c.font = lf
            if j in (2, 3, 11):
                c.alignment = ctr

    wb = Workbook(); wb.remove(wb.active)
    allws = wb.create_sheet('全部线索'); hdrst(allws)
    i = 2
    for c in codes:
        for row in merged[c]:
            wr(allws, i, row, c); i += 1
    allrows = i - 2

    for c in codes:
        ws = wb.create_sheet(f'{CN[c]}{c}'); hdrst(ws)
        for k, row in enumerate(merged[c], 2):
            wr(ws, k, row, c)

    sm = wb.create_sheet('汇总', 0)
    sm['A1'] = title; sm['A1'].font = Font(name='Arial', bold=True, size=14)
    sm.merge_cells('A1:F1')
    for j, h in enumerate(['国家', '代码', '地区', '线索数', '含邮箱', '含电话'], 1):
        c = sm.cell(3, j, h); c.font = navy; c.fill = hdr; c.border = bd
        c.alignment = Alignment(horizontal='center')
    r = 4
    for c in codes:
        sm.cell(r, 1, CN[c]); sm.cell(r, 2, c); sm.cell(r, 3, REG[c])
        sm.cell(r, 4, f'=COUNTIF(全部线索!$B$2:$B${allrows+1},"{CN[c]}")')
        sm.cell(r, 5, f'=COUNTIFS(全部线索!$B$2:$B${allrows+1},"{CN[c]}",全部线索!$H$2:$H${allrows+1},"?*")')
        sm.cell(r, 6, f'=COUNTIFS(全部线索!$B$2:$B${allrows+1},"{CN[c]}",全部线索!$I$2:$I${allrows+1},"?*")')
        for j in range(1, 7):
            cc = sm.cell(r, j); cc.border = bd; cc.font = cf
            if j >= 4:
                cc.alignment = Alignment(horizontal='center')
        r += 1
    sm.cell(r, 1, '合计').font = Font(name='Arial', bold=True)
    for col, l in [(4, 'D'), (5, 'E'), (6, 'F')]:
        cc = sm.cell(r, col, f'=SUM({l}4:{l}{r-1})')
        cc.font = Font(name='Arial', bold=True); cc.border = bd
        cc.alignment = Alignment(horizontal='center')
    for j in (1, 2, 3):
        sm.cell(r, j).border = bd
    for j, w in enumerate([14, 8, 12, 10, 10, 10], 1):
        sm.column_dimensions[get_column_letter(j)].width = w

    nt = wb.create_sheet('说明')
    total = sum(len(v) for v in merged.values())
    em = sum(1 for v in merged.values() for x in v if x.get('email'))
    ph = sum(1 for v in merged.values() for x in v if x.get('phone'))
    notes = [
        ['数据说明 / 合规', ''],
        ['来源', '联网公开检索的真实企业：本地企业名录与企业官网。每条含信息来源链接，可逐条核验。'],
        ['合规', '仅收录公开商务联系方式；未公开的邮箱/电话留空，不猜测、不编造。'],
        ['覆盖', f'{len(codes)} 国 {total} 家：含公开邮箱 {em}、含公开电话 {ph}。'],
        ['置信度', '高=官网/一手页面确认；中=来自权威名录，建议二次核验。'],
    ]
    for i2, (a, b) in enumerate(notes, 1):
        nt.cell(i2, 1, a).font = Font(name='Arial', bold=True, size=12 if i2 == 1 else 10)
        cb = nt.cell(i2, 2, b); cb.font = Font(name='Arial', size=10)
        cb.alignment = Alignment(wrap_text=True, vertical='top')
    nt.column_dimensions['A'].width = 14
    nt.column_dimensions['B'].width = 95

    order = ['汇总', '全部线索'] + [f'{CN[c]}{c}' for c in codes] + ['说明']
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
    wb.save(out_path)
    print(f'saved {out_path}: {total} rows, {len(codes)} countries, {em} emails, {ph} phones')


if __name__ == '__main__':
    main()
